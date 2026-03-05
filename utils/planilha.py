from pathlib import Path
from openpyxl import load_workbook


class Planilha:
    def __init__(self, planilha):
        self.__planilha = Path(planilha)
        self.wb = None
        self.ws = None
        self.cfg = None

    # Abrir planilha
    def recebendo_planilha(self):
        self.wb = load_workbook(self.__planilha)
        self.ws = self.wb.active
        # Criar coluna STATUS se não existir
        if self.ws.cell(row=1, column=7).value != "STATUS":
            self.ws.cell(row=1, column=7).value = "STATUS"


    # Buscar email pelo CPF
    def buscar_email_por_cpf(self, cpf_beneficiario):

        if not self.ws:
            self.recebendo_planilha()

        encontrado = False
        email_encontrado = None

        for row in self.ws.iter_rows(min_row=2):

            cpf_planilha = str(row[4].value)  # coluna CPF
            email = row[5].value              # coluna Email

            if cpf_planilha == str(cpf_beneficiario):
                email_encontrado = email
                row[6].value = "ENCONTRADO"
                encontrado = True

        if not encontrado:
            return False

        # salvar alterações
        self.wb.save(self.__planilha)

        return email_encontrado